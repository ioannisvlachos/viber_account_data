import json
import openpyxl
import os
import easygui

def options_menu(i):
	while i != '1' and i != '2':
		print('[*] Press 1 to view data')
		print('[*] Press 2 to export data')
		i = input('')
		clear_console()
	return i

def choose_cat(dictionary):
	c1 = 1
	c2 = 1
	list1 = []
	list2 = []
	print('\nData is presented in two categories, Main Category and Sub Category.')
	print('Firstly, select option from Main Category and then select option from Sub Category.')
	print('\nMain Category'+'\t'+'Sub Category\n')
	for keys in dictionary.keys():
		print('{}. {}'.format(c1, keys))
		list1.append(keys)
		c1 += 1
		for items in dictionary[keys]:
			print('{}{}. {}'.format('\t\t', c2, items))
			list2.append(items)
			c2 += 1
		print('\n')
	return list1, list2

def make_selection():
	main_input = int(input('Select Main Category: '))
	sub_input = int(input('Select Sub Category: '))
	return main_input, sub_input

def return_result(list1, list2, main_input, sub_input, dictionary):
	con = dictionary[list1[main_input-1]][list2[sub_input-1]]
	clear_console()
	print(list2[sub_input-1])
	print('\n')	
	#print(type(con))
	if type(con) is str:
		print(con)
		print('\n')
		#print(type(con))
	if type(con) is list:
		#print(type(con))
		for x in con:
			#print(type(x))
			if type(x) is dict:
				for y in x.items():
					print(y)
				print('\n')
			if type(x) is not dict:		
				print(x)
				print('\n')
def clear_console():
	if os.name == 'posix':
		os.system('clear')
	else:
		os.system('cls')	

def load_json():
	file = open(easygui.fileopenbox(title = 'Viber', msg = 'Select JSON file', default = '*.json'), encoding = 'utf8')
	data = json.load(file)
	return data

def create_sheet(data, wb):
	for key1 in data.keys():
		for key2 in data[key1]:
			sheet = wb.create_sheet(key2)

def remove_first_sheet(wb):
	sheet = wb.worksheets[0]
	wb.remove(sheet)		

def export_data(data, wb, index_row):
	index_row = 1
	for key1 in data.keys():
		for key2 in data[key1]:
			print(key2)
			typeof = type(data[key1][key2])
			if key2 == 'Email':
				key2 = 'Email1'
				print('in here')
			if key2 == 'Devices':
				key2 = 'Devices1'	
			sheet = wb[key2]
			if typeof is str:
				if key2 == 'Email1':
					key2 = 'Email'
				if key2 == 'Devices1':
					key2 = 'Devices'	
				sheet.cell(row = index_row, column = 1).value = str(data[key1][key2])
			if typeof is list:
				for x in data[key1][key2]:
					if type(x) is dict:
						for items in x.items():
							sheet.cell(row = index_row, column = 1).value = str(items[0])
							sheet.cell(row = index_row, column = 2).value = str(items[1])
							index_row += 1	
						index_row += 1	
					if type(x) is str:
						sheet.cell(row = index_row, column = 1).value = str(x)
						index_row += 1	
				index_row = 1	
	wb.save('export.xlsx')


#variables
c1 = 1
c2 = 1
list1 = []
list2 = []
i = '0'
index_row = 1


#main
data = load_json()
while i == '0':
	clear_console()
	i = options_menu(i)
	if i == '1':
		try:
			lists = choose_cat(data)
			list1 = lists[0]
			list2 = lists[1]
			selections = make_selection()
			main_input = selections[0]
			sub_input = selections[1]
			return_result(list1, list2, main_input, sub_input, data)
			i = input('Press 0 to continue, any key to exit..')
			clear_console()
		except Exception:
			clear_console()
			i = '0'
			print('\n[ERROR] Try again with different input..\n')
		
	if i == '2':
		wb = openpyxl.Workbook()
		create_sheet(data, wb)
		remove_first_sheet(wb)
		export_data(data, wb, index_row)
		clear_console()
		print('\n[*] Data exported successfully\n')
		i = input('Press 0 to continue, any key to exit..')
		clear_console()



